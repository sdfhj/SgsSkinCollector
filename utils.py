import os
import requests
from faker import Faker
faker = Faker(locale='zh_CN')
from openpyxl import load_workbook


def mkdir(path):
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)

def get(url):
    header = {'User-Agent': faker.user_agent()}
    i = 0
    while i < 1:
        try:
            result = requests.get(url, headers=header, timeout=10)
            return result
        except requests.exceptions.RequestException:
            # print("Time out " + str(i+1))
            i += 1

def dld(path, name, url, debug = False):
    try:
        if not os.path.isfile(path + '/' + name):
            if debug:
                print('File does not exist')
            document = get(url)
            if document.status_code == 200:
                mkdir(path)
                with open(path + '/' + name, 'wb') as f:
                    f.write(document.content)
                    if debug:
                        print('File download completed')
                    return True
            else:
                if debug:
                    print('File download failed and ignored')
                return False
        else:
            if debug:
                print('File exists')
            return True
    except Exception:
        if debug:
            print('File download failed and ignored')
        return False


def get_sort_info(local_path): 
    index_2_list = []
    name_list = []
    wb = load_workbook(filename = local_path + '/' + 'sort.xlsx')
    ws = wb['Sheet1']
    rows = ws.rows
    for row in rows:
        line = [col.value for col in row]
        
        if line[2] == None:
            name = line[3]
        else:
            name = line[2] + '·' + line[3]
        name = str(name)
        name = str(line[0]).zfill(4) + ' - ' + name.replace(u'\u3000',u'')

        index_2_list.append(str(line[0]).zfill(4))
        name_list.append(name)

    return name_list, index_2_list

def get_sort_info_female(local_path): 
    index_2_list = []
    name_list = []
    wb = load_workbook(filename = local_path + '/' + 'sort.xlsx')
    ws = wb['Sheet1']
    rows = ws.rows
    for row in rows:
        line = [col.value for col in row]
        
        if line[2] == None:
            name = line[3]
        else:
            name = line[2] + '·' + line[3]
        name = str(name)
        name = str(line[0]).zfill(4) + ' - ' + name.replace(u'\u3000',u'')

        if line[4] == '女':
            index_2_list.append(str(line[0]).zfill(4))
            name_list.append(name)

    return name_list, index_2_list

def get_example_info(local_path): 
    url_list = []
    folder_list = []
    file_name_list = []
    wb = load_workbook(filename = local_path + '/' + 'Example.xlsx')
    ws = wb['Sheet1']
    rows = ws.rows
    for row in rows:
        line = [col.value for col in row]

        url_list.append(line[2])
        folder_list.append(line[3])
        file_name_list.append(line[4])

    return url_list[1:], folder_list[1:], file_name_list[1:]

def print_progress(finished, all, time_dur):
    i = finished / all * 100
    # print("\r\t\t%.2f %%:" % (i), '|', "▋" * (int(i) // 2),"-" * (50 - (int(i) // 2)), '|', '%.2f' % time, 's', end="")
    print("\r   {:.2f} % |{}{}| {:.2f}s".format(i, "▋" * (int(i) // 2), "-" * (50 - (int(i) // 2)), time_dur), end="")