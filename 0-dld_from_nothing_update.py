import os
import requests
from faker import Faker
faker = Faker(locale='zh_CN')
import threading
from openpyxl import load_workbook
import time
import shutil


def mkdir(path):
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)

def get(url):
    header = {'User-Agent': faker.user_agent()}
    i = 0
    while i < 1:
        try:
            result = requests.get(url, headers=header, timeout=10)
            return result
        except requests.exceptions.RequestException:
            # print("Time out " + str(i+1))
            i += 1

def dld(path, name, url):
    try:
        if not os.path.isfile(path + '/' + name):
            # print('不存在')
            document = get(url)
            if document.status_code == 200:
                mkdir(path)
                with open(path + '/' + name, 'wb') as f:
                    f.write(document.content)
                    # print('已下载完成')
                    return True
            else:
                # print('下载失败并忽略')
                return False
        else:
            # print('已存在')
            return True
    except Exception:
            # print('下载失败并忽略')
            return False

def dld_index_2(index_2):
    Semaphore.acquire()
    index_2_4 = index_2.zfill(4)
    # print(index_2)
    old_dld_path = local_path_sort + '/' + index_2_4
    if index_2_4 in index_2_list:
        new_dld_path = local_path_sort + '/' + str(name_list[index_2_list.index(index_2_4)])
        dld_path = new_dld_path
        if os.path.exists(old_dld_path):
            os.rename(old_dld_path, new_dld_path)
    else:
        dld_path = old_dld_path
    
    for index_1 in range(1,7):
        # index_1: quality of the skin [1, 6]
        # print('\t' + str(index_1))
        indicator = 0
        for index_3 in range(1,10):
            skin_number = str(index_1) + index_2 + str(index_3).zfill(2)
            # for dynamic skin on pc index_1 plus 2
            skin_number_plus2 = str(index_1 + 2) + index_2 + str(index_3).zfill(2)

            # try to download default type of file
            if folder_list[0] is not None:
                file_path = dld_path + '/' + folder_list[0]
            else:
                file_path = dld_path + '/'
            if not dld(file_path.replace('skin_number_plus2', skin_number_plus2).replace('skin_number', skin_number), file_name_list[0].replace('skin_number_plus2', skin_number_plus2).replace('skin_number', skin_number), url_list[0].replace('skin_number_plus2', skin_number_plus2).replace('skin_number', skin_number)):
                # if the default type of file does not exist, indicator ++
                indicator = indicator + 1
            else:
                indicator = 0
            
            # if the 1bbbb01 file does not exist
            if (indicator > 0) & (index_1 == 1) & (index_3 == 1):
                # index_2 is not valid, return None and try another index_2
                Semaphore.release()
                return None

            # if the abbbb01 file does not exist
            # if (indicator > 0) & (index_3 == 1):
            #     # index_1 is not valid, try next index_1
            #     break

            # if the abbbbcc file does not exist
            # if (indicator > 0):
            #     # index_3 is not valid, try next index_3
            #     continue

            # if indicator above certain value
            if (indicator > 2):
                break

            # try to download the rest types of the files
            for file_type_index, file_name in enumerate(file_name_list[1:]):
                if folder_list[file_type_index + 1] is not None:
                    file_path = dld_path + '/' + folder_list[file_type_index + 1]
                else:
                    file_path = dld_path + '/'
                url = url_list[file_type_index + 1]
                dld(file_path.replace('skin_number_plus2', skin_number_plus2).replace('skin_number', skin_number), file_name.replace('skin_number_plus2', skin_number_plus2).replace('skin_number', skin_number), url.replace('skin_number_plus2', skin_number_plus2).replace('skin_number', skin_number))

    Semaphore.release()

def get_sort_info(local_path): 
    index_2_list = []
    name_list = []
    wb = load_workbook(filename = local_path + '/' + 'sort.xlsx')
    ws = wb['Sheet1']
    rows = ws.rows
    for row in rows:
        line = [col.value for col in row]
        
        if line[2] == None:
            name = line[3]
        else:
            name = line[2] + '·' + line[3]
        name = str(name)
        name = str(line[0]).zfill(4) + ' - ' + name.replace(u'\u3000',u'')

        index_2_list.append(str(line[0]).zfill(4))
        name_list.append(name)

    return name_list, index_2_list

def get_example_info(local_path): 
    url_list = []
    folder_list = []
    file_name_list = []
    wb = load_workbook(filename = local_path + '/' + 'Example.xlsx')
    ws = wb['Sheet1']
    rows = ws.rows
    for row in rows:
        line = [col.value for col in row]

        url_list.append(line[2])
        folder_list.append(line[3])
        file_name_list.append(line[4])

    return url_list[1:], folder_list[1:], file_name_list[1:]

def print_progress(finished, all, time_dur):
    i = finished / all * 100
    # print("\r\t\t%.2f %%:" % (i), '|', "▋" * (int(i) // 2),"-" * (50 - (int(i) // 2)), '|', '%.2f' % time, 's', end="")
    print("\r   {:.2f} % |{}{}| {:.2f}s".format(i, "▋" * (int(i) // 2), "-" * (50 - (int(i) // 2)), time_dur), end="")

# -------------------------------------------------------------------------------------

# Settings
# Download path
local_path = "D:\\Another\\temp\\ANOTHER_RUBBISH\\VIDEOS\\sgsnew"
# local_path = 'C:\\下载\\temp\\sgs'
# Number of download threads
max_connections = 12

# Create the sort folder
local_path_sort = local_path + '\\sort'
mkdir(local_path_sort)
# Create a Bounded Semaphore
Semaphore = threading.BoundedSemaphore(max_connections)

# Get infomation about characters and file samples to be downloaded
(name_list, index_2_list) = get_sort_info(local_path)
(url_list, folder_list, file_name_list) = get_example_info(local_path)

# dld_index_2('031')

# Define a list, where all download threads are put into
threads = []
# Put all threads into the list
index_2_start = 1
index_2_end = 10000
index_2_len = index_2_end - index_2_start
for index_2 in range(index_2_start, index_2_end):
    threads.append(threading.Thread(target=dld_index_2, args=(str(int(index_2)).zfill(3),)))
# Use start() method to start the threads
for thread in threads:
    thread.start()
# Use join() method to let the main thread waiting for the finish of subthreads
# for thread in threads:
#     thread.join()

# Progress Bar
print('Download Progress:')
start_time = time.perf_counter()
while len(threading.enumerate()) > 1:
    time.sleep(0.5)
    current_time = time.perf_counter()
    print_progress(index_2_len - (len(threading.enumerate()) - 1), index_2_len, current_time - start_time)
current_time = time.perf_counter()
print_progress(index_2_len - (len(threading.enumerate()) - 1), index_2_len, current_time - start_time)
print('\n')
